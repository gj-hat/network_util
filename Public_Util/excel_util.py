"""
@author: JiaGuo
@email:
@date: 2024/08/19
@description: Excel 工具类，统一处理单元格内容，并提供多种读取和写入方式。
@version: 1.0
"""

import logging
from typing import Any, List, Optional, Union

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment

logger = logging.getLogger('ExcelUtil')


class ExcelUtil:
    """
    Excel 工具类，用于读取和写入 Excel 文件，同时对单元格内容进行统一处理。
    """

    def __init__(self, file_path: str) -> None:
        """
        初始化 ExcelUtil 实例

        :param file_path: Excel 文件路径
        """
        self.file_path = file_path
        self._workbook = None

    @property
    def workbook(self) -> openpyxl.Workbook:
        """
        懒加载 Excel 工作簿

        :return: openpyxl.Workbook 对象
        """
        if not self._workbook:
            self._workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        return self._workbook

    def process_cell(self, cell: Any) -> Any:
        """
        处理单元格内容，支持空值和多行文本  这里可以重写

        :param cell: 原始单元格内容（任意类型）
        :return: 处理后的内容。如果是空值返回空列表，字符串按换行符拆分为列表，否则原样返回
        """
        # if cell is None or (isinstance(cell, str) and cell.strip() == ''):
        #     return []
        # if isinstance(cell, str):
        #     cleaned = cell.strip('\n')
        #     return cleaned.split('\n') if '\n' in cleaned else [cleaned]
        return cell


    def write_excel(self, data: pd.DataFrame, sheet_name: str) -> None:
        """
        将 DataFrame 数据写入 Excel 文件

        :param data: 要写入的数据，类型为 DataFrame
        :param sheet_name: 写入的工作表名称
        """
        try:
            with pd.ExcelWriter(self.file_path) as writer:
                data.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            logger.error(f"写入 Excel 文件失败: {e}")

    from openpyxl.styles import Alignment

    def write_excel_cell(self, sheet_name: str, row: int, column_name: str, value: any) -> None:
        """
        写入 Excel 单元格数据（支持换行显示）

        :param sheet_name: 工作表名称
        :param row: 行号（从 1 开始）
        :param column_name: 列的字段名
        :param value: 要写入的值（支持多行）
        """
        try:
            if row is None:
                raise ValueError(f"写入失败：行号 row 不能为 None。列: {column_name}, 值: {value}")

            sheet = self.workbook[sheet_name]
            # 获取表头（第一行）
            header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))

            # 查找列名对应的列号
            if column_name in header:
                col_num = header.index(column_name) + 1
            else:
                raise ValueError(f"列名 '{column_name}' 在工作表 '{sheet_name}' 的表头中未找到。")

            # 写入数据并设置自动换行
            cell = sheet.cell(row=row, column=col_num, value=value)
            cell.alignment = Alignment(wrap_text=True)

            # 保存工作簿
            self.workbook.save(self.file_path)
            logger.info(f"成功在工作表 '{sheet_name}' 的第 {row} 行，列 '{column_name}' 写入数据（已设置换行）")
        except KeyError:
            logger.error(f"工作表 '{sheet_name}' 不存在")
        except Exception as e:
            logger.error(f"写入 Excel 单元格失败: {e}")

    #################  读取 ####################################
    def read_excel_df(self,
                      sheet_name: Optional[str] = None,
                      header: int = 0,
                      skiprows: Optional[Union[int, List[int]]] = None) -> Optional[pd.DataFrame]:
        """
        读取 Excel 文件并返回 DataFrame，对每个单元格调用 process_cell() 进行处理

        :param sheet_name: 工作表名称，默认为 None（读取所有工作表）
        :param header: 标题行索引（默认 0）
        :param skiprows: 跳过的行数或行号列表
        :return: 处理后的 DataFrame，读取失败返回 None
        """
        try:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=header, skiprows=skiprows)
            df = df.applymap(self.process_cell)
            return df
        except Exception as e:
            logger.error(f"读取 Excel 文件失败: {e}")
            return None

    def read_excel_dict(self,
                        sheet_name: Optional[str] = None,
                        header: int = 0,
                        row_num: Optional[int] = None,
                        skiprows: Optional[Union[int, List[int]]] = None) -> List[dict]:
        """
        读取 Excel 文件并转换为字典列表，每一行对应一个字典。
        如果 row_num 提供，则仅返回该行数据；否则返回所有数据。
        并在每个字典中添加 'id' 字段表示 Excel 中的真实行号（从1开始）
        """
        df = self.read_excel_df(sheet_name=sheet_name, header=header, skiprows=skiprows)
        if df is None:
            return []

        # 记录真实行号（从1开始）
        original_index = df.index + 1
        if row_num is not None:
            try:
                df = df.iloc[[row_num - 1]]
                row_ids = [row_num]
            except IndexError as e:
                logger.error(f"行号 {row_num} 超出范围: {e}")
                return []
        else:
            row_ids = list(original_index)

        records = df.to_dict(orient='records')
        for i, record in enumerate(records):
            record['row_id'] = row_ids[i]
        return records


    def read_excel_dict_by_key(self,
                               sheet_name: str,
                               key_column: str,
                               header: int = 0,
                               skiprows: Optional[Union[int, List[int]]] = None) -> dict:
        """
        读取 Excel 文件并转换为以指定列值为键的字典

        :param sheet_name: 工作表名称
        :param key_column: 用作键的列名
        :param header: 标题行索引（默认 0）
        :param skiprows: 跳过的行数或行号列表
        :return: 键为指定列值，值为对应行字典的字典；失败返回空字典
        """
        df = self.read_excel_df(sheet_name=sheet_name, header=header, skiprows=skiprows)
        if df is None:
            return {}
        if key_column not in df.columns:
            logger.error(f"指定的列 '{key_column}' 在工作表中不存在")
            return {}
        try:
            return df.set_index(key_column).T.to_dict()
        except Exception as e:
            logger.error(f"转换 DataFrame 为字典失败: {e}")
            return {}

    def filter_excel_by_column(self,
                               sheet_name: str,
                               column_name: str,
                               value: Any,
                               header: int = 0,
                               skiprows: Optional[Union[int, List[int]]] = None) -> List[dict]:
        """
        根据指定列的值筛选数据，返回字典列表
        :param sheet_name: 工作表名称
        :param column_name: 要筛选的列名
        :param value: 筛选的值
        :param header: 标题行索引（默认 0）
        :param skiprows: 跳过的行数或行号列表
        :return: 符合条件的字典列表，筛选失败返回空列表
        """
        df = self.read_excel_df(sheet_name=sheet_name, header=header, skiprows=skiprows)
        if df is None:
            return []
        if column_name not in df.columns:
            logger.error(f"指定的列 '{column_name}' 在工作表中不存在")
            return []
        try:
            filtered_df = df[df[column_name] == value]
            return filtered_df.to_dict(orient='records')
        except Exception as e:
            logger.error(f"根据列 '{column_name}' 筛选数据失败: {e}")
            return []


    def read_excel_data(self,
                        sheet_name: str,
                        row_num: Optional[int] = None,
                        col_num: Optional[int] = None) -> Union[Any, List[Any], List[List[Any]]]:
        """
        读取 Excel 数据，统一通过 read_excel_df() 读取并处理每个单元格，
        根据参数返回单个单元格、整行、整列或完整二维列表。

        :param sheet_name: 工作表名称
        :param row_num: 行号（从 1 开始，可选）
        :param col_num: 列号（从 1 开始，可选）
        :return: 单个值、行列表、列列表或完整二维列表
        """
        try:
            df = self.read_excel_df(sheet_name=sheet_name)
            if df is None:
                return None

            # 同时指定行和列：返回单个单元格处理后的值（注意 DataFrame 的索引从0开始）
            if row_num is not None and col_num is not None:
                return df.iat[row_num - 1, col_num - 1]

            # 仅指定行：返回该行所有单元格处理后的值
            if row_num is not None:
                return df.iloc[row_num - 1].tolist()

            # 仅指定列：返回该列所有单元格处理后的值
            if col_num is not None:
                return df.iloc[:, col_num - 1].tolist()

            # 未指定行或列：返回整个工作表的二维数据
            return df.values.tolist()

        except IndexError as e:
            logger.error(f"索引错误: {e}")
            return None if (row_num is not None or col_num is not None) else []
        except Exception as e:
            logger.error(f"读取 Excel 数据时出错: {e}")
            return None if (row_num is not None or col_num is not None) else []


    def get_header(self,
                   sheet_name: str,
                   header: int = 0,
                   skiprows: Optional[Union[int, List[int]]] = None) -> List[Any]:
        """
        获取指定工作表的 header 数据，经过 process_cell() 处理后返回

        :param sheet_name: 工作表名称
        :param header: 标题行索引（默认 0）
        :param skiprows: 跳过的行数或行号列表
        :return: 经过处理的 header 列表
        """
        try:
            # 通过 pd.read_excel() 仅读取 header 行（nrows=0）
            df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=header, skiprows=skiprows, nrows=0)
            # df.columns 获取的是 header 列表，对每个单元格调用 process_cell() 进行处理
            return [self.process_cell(col) for col in list(df.columns)]
        except Exception as e:
            logger.error(f"获取 header 失败: {e}")
            return []


if __name__ == '__main__':
    # 示例用法
    excel_file = "data.xlsx"
    excel_util = ExcelUtil(excel_file)

    # 读取工作表为 DataFrame
    df = excel_util.read_excel_df(sheet_name="device_cmd_often")
    print("DataFrame 数据:", df)

    # 读取工作表为字典列表
    dict_list = excel_util.read_excel_dict(sheet_name="device_cmd_often")
    print("字典列表:", dict_list)

    # 根据指定列作为键读取数据
    dict_by_key = excel_util.read_excel_dict_by_key(sheet_name="device_cmd_often", key_column="设备类型")
    print("指定列作为键的字典:", dict_by_key)

    # 根据指定列值筛选数据
    filtered = excel_util.filter_excel_by_column(sheet_name="device_cmd_often", column_name="设备类型", value="h3c")
    print("筛选后的数据:", filtered)

    # 读取单元格、行、列和整个工作表数据
    single_value = excel_util.read_excel_data(sheet_name="device_cmd_often", row_num=2, col_num=3)
    print("单元格数据:", single_value)
    row_data = excel_util.read_excel_data(sheet_name="device_cmd_often", row_num=2)
    print("行数据:", row_data)
    col_data = excel_util.read_excel_data(sheet_name="device_cmd_often", col_num=2)
    print("列数据:", col_data)
    full_data = excel_util.read_excel_data(sheet_name="device_cmd_often")
    print("完整工作表数据:", full_data)
